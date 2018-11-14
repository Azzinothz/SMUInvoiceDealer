from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment


logs = []

with open("./log.txt", "r", encoding="UTF-8") as f:
    lines = f.readlines()

for line in lines:
    line = line.rstrip("\n")
    log = {
        "member": line[:3],
        "activity": line[-4:]
    }
    if line[13] != "日":
        log["date"] = line[4: 13]
        line = line[15:]
    else:
        log["date"] = line[4: 14]
        line = line[16:]
    log["destination"] = line[: -6]
    logs.append(log)

wb = load_workbook("SummaryTemplate.xlsx")
ws = wb.active
left, right, top, bottom = [Side(style='thin', color='000000')]*4
count = 3
for log in logs:
    ws["A" + str(count)].value, ws["B" + str(count)], ws["C" + str(count)], ws["D" + str(count)] = \
        log["date"], log["destination"], log["activity"], log["member"]
    row = ws["A" + str(count) + ":D" + str(count)][0]
    for cell in row:
        cell.border = Border(left, right, top, bottom)
        cell.alignment = Alignment(horizontal='center')
    count += 1
wb.save("Summary.xlsx")

print("============================================================")
print("COUNT:", len(logs))
print("============================================================")
