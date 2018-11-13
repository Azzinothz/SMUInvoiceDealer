import random

from openpyxl import load_workbook
from openpyxl.styles import Border, Side


data = {
    "name": ["尤逸昊", "傅午辰", "蔡俊弛"],
    "start_date": [2018, 2, 1],
    "for": ["问卷调查", "路演宣传", "发放问卷", "录制视频", "素材拍摄", "随机采访"],
    "price": {
        4: 0,
        5: 0,
        6: 32,
        7: 8,
        8: 88,
        9: 200,
        10: 56,
        11: 0,
        12: 0,
    },
    "route": {
        4: "书院",
        5: "惠南",
        6: "南六公路",
        7: "新场",
        8: "罗山路",
        9: "龙阳路",
        10: "世纪大道",
        11: "金桥路",
        12: "高桥",
    },
    "destination": {
        4: ["书院镇", "泥城镇", "万祥镇"],
        5: ["浦东商城", "南汇汽车站", "浦东颐景园"],
        6: ["上海野生动物园", "欧尚", "上海两港装饰城"],
        7: ["新场镇", "新场大街", "新场购物广场"],
        8: ["宜家", "长泰国际", "万达"],
        9: ["麦德龙", "大唐国际", "永达国际", "喜马拉雅中心", "花木公园", "建平中学西校"],
        10: ["八佰伴", "时代广场", "世纪大道", "福山外国语小学"],
        11: ["金桥国际", "久金", "金桥站", "上海理工大学"],
        12: ["高桥中学", "高桥古镇", "纺织展示馆"]
    }
}


def update_date():
    if data["start_date"][2] + 1 == 29:
        data["start_date"][1] += 1
        data["start_date"][2] = 1
    else:
        data["start_date"][2] += 1


def fill_data(sheet):
    data["name"] += random.sample(data["name"], 1)
    cells = sheet["C5:I8"]
    total = 0
    desc = 14
    reasons = ()
    for i in range(len(cells)):
        cells[i][0].value = data["name"][i]
        if i == 3:
            update_date()
        cells[i][1].value = "{0}.{1}.{2}".format(data["start_date"][0], data["start_date"][1], data["start_date"][2])

        reason = random.sample(data["for"], 1)[0]
        cells[i][2].value = reason

        price = -1
        for w_amount in data["price"]:
            if data["price"][w_amount] > 1:
                price = w_amount
                data["price"][w_amount] -= 2
                break

        if reason not in reasons:
            reasons += (reason,)
            desc += 1
            sheet["C" + str(desc)] = "于{0}进行{1}".format(random.sample(data["destination"][price], 1)[0], reason)

        cells[i][3].value = "海事大学-" + data["route"][price]
        cells[i][4].value = str(price) + "×2"
        cells[i][6].value = str(price * 2)
        total += price * 2

    update_date()
    sheet["C9"] = "合计人民币(大写)：" + to_currency(total) + "整"
    sheet["I9"] = str(total)
    return total


def to_currency(number):
    if not isinstance(number, float) and not isinstance(number, int):
        return 'non number'
    if number < 0 or number > 9999999999999.99:
        return 'wrong number'
    if number == 0:
        return '零圆'
    c_d = {'0': '零', '1': '壹', '2': '贰', '3': '叁', '4': '肆', '5': '伍', '6': '陆', '7': '柒', '8': '捌', '9': '玖'}
    d_d = {0: '分', 1: '角', 2: '圆', 3: '拾', 4: '佰', 5: '仟', 6: '万', 7: '拾', 8: '佰', 9: '仟', 10: '亿', 11: '拾', 12: '佰',
           13: '仟',
           14: '万'}
    L = []
    pre = '0'
    s = str(int(number * 100))[::-1].replace('.', '')
    index = -1

    for c in s:
        index += 1
        if c == '0' and pre == '0':
            if index == 2:
                L.insert(0, '圆')
        elif c == '0':
            if index == 2:
                L.insert(0, '圆')
            else:
                L.insert(0, '零')
            pre = c
        else:
            L.insert(0, c_d[c] + "" + d_d[index])
            pre = c
    return ''.join(L)


def reformat(sheet):
    left, right, top, bottom = [Side(style='thin', color='000000')]*4
    cells = sheet["E10:I10"][0]
    for cell in cells:
        cell.border = Border(top=top, bottom=bottom)
    cells[0].border += Border(left=left)
    cells[-1].border += Border(right=right)


if __name__ == "__main__":
    wb = load_workbook("市内交通费报销单.xlsx")
    ws = wb.active
    count = 0
    for amount in data["price"]:
        if data["price"][amount] > 1:
            count += data["price"][amount]

    total_amount = 0
    err_page = []
    for page in range(count // 8):
        total_amount += fill_data(ws)
        reformat(ws)
        if len({ws["I5"].value, ws["I6"].value, ws["I7"].value, ws["I8"].value}) > 1:
            err_page.append(page)
        wb.save("./result/市内交通费报销单" + str(page) + ".xlsx")

    print("============================================================")
    print("TOTAL:", total_amount)
    print("PAGES:", len(range(count // 8)))
    print("ERROR:", err_page)
    print("============================================================")